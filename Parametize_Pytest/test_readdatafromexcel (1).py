from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import pytest
import openpyxl

class Test_SauceDemo:
    def setup_method(self):
        #her test başlangıcında çalışacak fonksiyon
        self.driver = webdriver.Chrome()
        self.driver.maximize_window() 
        self.driver.get("https://www.saucedemo.com/")

    def teardown_method(self):
        #her test bitiminde çalışacak fonksiyon
        self.driver.quit()

    def getData():
        excelFile = openpyxl.load_workbook("data/saucedemo.xlsx") #dosya açılıyor
        sheet = excelFile["Sayfa1"] #hangi sayfada çalıştığımızı belirtiyoruz
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):# saymaya 0 dan başlar.
            username = sheet.cell(i,1).value #i. satırın 1. hücresi(cell) username'i tutuyor
            password = sheet.cell(i,2).value
            data.append((username,password))
    
        return data
    
    @pytest.mark.parametrize("username,password", getData())
    def test_invalid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"user-name")))
        usernameInput.send_keys(username)
        passwordInput =WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"password")))
        passwordInput.send_keys(password)
        loginButton = WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"login-button")))
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
    